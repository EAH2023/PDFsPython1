import re
import pandas as pd
import pdfplumber
from openpyxl import load_workbook

def procesar_pdf(pdf_path):
    """
    Procesa un archivo PDF y extrae datos estructurados de cada página.
    Divide el campo Concepto en subconceptos con sus importes.
    """
    resultados = []
    
    # Abrir el PDF con pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            texto = page.extract_text()
            if not texto:
                continue
            lineas = texto.splitlines()
            
            # Extraer detalles de las líneas específicas
            fecha_vcto = re.search(r"\d{2}-\d{2}-\d{2}", lineas[2]).group(0) if len(lineas) > 2 and re.search(r"\d{2}-\d{2}-\d{2}", lineas[2]) else None
            importe = re.search(r"(\d+,\d+)$", lineas[13]).group(1) if len(lineas) > 13 and re.search(r"(\d+,\d+)$", lineas[13]) else None
            concepto = lineas[7].strip() if len(lineas) > 7 else None

            # Limpieza y formateo
            if concepto:
                concepto = concepto.replace("TORTOLA 4 ", "").strip()  # Eliminar texto superfluo
                # Dividir en subconceptos (texto seguido de un número decimal)
                subconceptos = re.findall(r"(.*?)(\d+,\d+)", concepto)
                for subconcepto, imp_subconcepto in subconceptos:
                    subconcepto = subconcepto.strip().rstrip(",")  # Limpiar texto y eliminar comas finales o iniciales
                    if subconcepto.startswith(","):
                        subconcepto = subconcepto[1:].strip()  # Quitar coma inicial si la hay
                    detalles = {
                        "Num. Recibo": re.search(r"EUR\s+(\d+)", lineas[2]).group(1) if len(lineas) > 2 and re.search(r"EUR\s+(\d+)", lineas[2]) else None,
                        "Concepto": subconcepto,
                        "Importe (€)": f"{float(importe.replace(',', '.')):.2f} €" if importe else None,
                        "Fecha Vcto": pd.to_datetime(fecha_vcto, format="%d-%m-%y").strftime("%d/%m/%y") if fecha_vcto else None,
                        "Imp.Subconcepto (€)": f"{float(imp_subconcepto.replace(',', '.')):.2f} €",  # Importe del subconcepto
                    }
                    resultados.append(detalles)

    return pd.DataFrame(resultados)

def guardar_en_excel_existente(dataframe, output_path):
    """
    Añade datos al archivo Excel existente en la primera hoja 'Apuntes'.
    Si no existe el archivo, lo crea.
    """
    try:
        # Cargar el archivo Excel existente
        book = load_workbook(output_path)
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # Determinar la posición inicial de los nuevos datos
            startrow = writer.sheets["Apuntes"].max_row if "Apuntes" in writer.sheets else 0
            dataframe.to_excel(writer, sheet_name="Apuntes", index=False, header=startrow == 0, startrow=startrow)
    except FileNotFoundError:
        # Crear un nuevo archivo Excel si no existe
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            dataframe.to_excel(writer, sheet_name="Apuntes", index=False)
    print(f"Datos añadidos al archivo Excel: {output_path}")

# Ruta del archivo PDF de entrada
pdf_path = "archivo2.pdf"  # Cambia esto por la ruta de tu archivo PDF

# Ruta del archivo Excel de salida
output_excel = "datos_extraidos.xlsx"

# Procesar el PDF y guardar los resultados en el Excel existente
df_resultados = procesar_pdf(pdf_path)
guardar_en_excel_existente(df_resultados, output_excel)
